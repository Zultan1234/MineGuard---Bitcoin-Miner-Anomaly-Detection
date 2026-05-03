"""
Miner Data Collector — Standalone 24/7 Poller
==============================================
Polls ALL cgminer API commands from an Antminer L3+ every 5 seconds.
Extracts EVERY field from EVERY response section, including:
  - summary  : hashrate, accepted/rejected, HW errors, uptime, pool latency
  - stats    : per-chain hashrate, temps, fans, freq, voltage, chip counts
  - devs     : per-board device stats (one row prefix per board)
  - pools    : per-pool stats (accepted, rejected, latency, difficulty)

Multi-item sections (devs has 4 boards, pools has N pools) are
flattened with prefixes: dev0_Accepted, dev1_Accepted, pool0_URL, etc.

Usage:
    python miner_collector.py --ip 192.168.1.100 --port 4028
"""

import socket, json, re, time, os, sys, argparse, threading
from datetime import datetime

try:
    from colorama import init as colorama_init, Fore, Style
    colorama_init()
    C_GREEN=Fore.GREEN; C_YELLOW=Fore.YELLOW; C_RED=Fore.RED
    C_CYAN=Fore.CYAN; C_RESET=Style.RESET_ALL; C_BOLD=Style.BRIGHT
except ImportError:
    C_GREEN=C_YELLOW=C_RED=C_CYAN=C_RESET=C_BOLD=""

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

POLL_INTERVAL = 5
BACKUP_FILE   = "miner_backup.json"
EXCEL_DIR     = "exports"
TIMEOUT       = 10

ALWAYS_SKIP = {
    "id","When","Code","Elapsed","Description","Msg","STATUS",
    "CGMiner","Miner","CompileTime","Type","ID","STATS",
    "chain_acs1","chain_acs2","chain_acs3","chain_acs4",
    "chain_acs5","chain_acs6",
    "Min","Max","Wait","Calls","fan_num","temp_num",
}

# ── TCP ──────────────────────────────────────────────────────────────────────

def _tcp_query(ip, port, command):
    raw = b""
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.settimeout(TIMEOUT)
        s.connect((ip, port))
        s.sendall(json.dumps({"command": command}).encode())
        time.sleep(0.2)
        while True:
            try:
                chunk = s.recv(65536)
                if not chunk: break
                raw += chunk
            except socket.timeout:
                break
    text = raw.decode(errors="replace").replace("\x00","").strip()
    if not text:
        raise ConnectionError(f"No data for '{command}'")
    text = re.sub(r'}\s*{', '},{', text)
    return json.loads(text)

# ── FIELD EXTRACTION ─────────────────────────────────────────────────────────

def _try_float(v):
    if isinstance(v, bool):         return None
    if isinstance(v, (int,float)):  return float(v)
    if isinstance(v, str):
        try:    return float(v.strip())
        except: return None
    return None

def _is_meta(item):
    return "CGMiner" in item or "CompileTime" in item or "Type" in item

def _flatten(item, prefix=""):
    out = {}
    for k, v in item.items():
        if k in ALWAYS_SKIP: continue
        col = f"{prefix}{k}" if prefix else k
        fv = _try_float(v)
        if fv is not None:
            out[col] = fv
        elif isinstance(v, str) and len(v) < 120:
            out[col] = v
    return out

def extract_summary(raw):
    items = raw.get("SUMMARY", [])
    return _flatten(items[0]) if items else {}

def extract_stats(raw):
    out = {}
    for item in raw.get("STATS", []):
        if isinstance(item, dict) and not _is_meta(item):
            out.update(_flatten(item))
    return out

def extract_devs(raw):
    out = {}
    for i, item in enumerate(raw.get("DEVS", [])):
        if isinstance(item, dict):
            out.update(_flatten(item, prefix=f"dev{i}_"))
    return out

def extract_pools(raw):
    out = {}
    for i, item in enumerate(raw.get("POOLS", [])):
        if isinstance(item, dict):
            out.update(_flatten(item, prefix=f"pool{i}_"))
    return out

def poll_all(ip, port):
    jobs = [
        ("summary", extract_summary),
        ("stats",   extract_stats),
        ("devs",    extract_devs),
        ("pools",   extract_pools),
    ]
    merged, errors = {}, {}
    for cmd, extractor in jobs:
        try:
            merged.update(extractor(_tcp_query(ip, port, cmd)))
        except Exception as e:
            errors[cmd] = str(e)
    merged["_errors"] = errors
    return merged

# ── BACKUP ───────────────────────────────────────────────────────────────────

def load_backup(path):
    records = []
    if not os.path.exists(path): return records
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line: continue
            try: records.append(json.loads(line))
            except: pass
    return records

def append_backup(path, record):
    with open(path, "a", encoding="utf-8") as f:
        f.write(json.dumps(record, ensure_ascii=False) + "\n")

# ── EXCEL EXPORT ─────────────────────────────────────────────────────────────

def section_order(k):
    if k == "timestamp":             return (0, k)
    if k.startswith("dev"):          return (3, k)
    if k.startswith("pool"):         return (4, k)
    stats_kw = {"temp","fan","chain","freq","volt","rate","chip","total","ghs","mhs"}
    if any(x in k.lower() for x in stats_kw): return (2, k)
    return (1, k)

SECTION_COLORS = {0:"1F3864", 1:"1F3864", 2:"1B5E20", 3:"4A235A", 4:"7B341E"}
SECTION_NAMES  = {0:"timestamp", 1:"summary", 2:"stats", 3:"devs", 4:"pools"}

def export_excel(records, label=""):
    if not EXCEL_OK:
        raise RuntimeError(
            "openpyxl not installed.\n"
            "Run: C:\\Users\\User\\AppData\\Local\\Programs\\Python\\Python313\\"
            "python.exe -m pip install openpyxl"
        )
    if not records:
        raise ValueError("No data collected yet.")

    os.makedirs(EXCEL_DIR, exist_ok=True)

    all_keys = set()
    for r in records:
        all_keys.update(k for k in r if not k.startswith("_"))

    sorted_cols = sorted(all_keys, key=section_order)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Miner Data"

    hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="Arial", size=9)
    ts_font   = Font(name="Courier New", size=9)
    alt_fill  = PatternFill("solid", start_color="EEF2FF")
    c_align   = Alignment(horizontal="center")
    thin      = Side(style="thin", color="CCCCCC")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.row_dimensions[1].height = 36
    for ci, col in enumerate(sorted_cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font      = hdr_font
        cell.alignment = hdr_align
        cell.border    = border
        so = section_order(col)[0]
        cell.fill = PatternFill("solid", start_color=SECTION_COLORS.get(so,"1F3864"))

    for ri, rec in enumerate(records, 2):
        is_alt = (ri % 2 == 0)
        for ci, col in enumerate(sorted_cols, 1):
            val  = rec.get(col, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = border
            if is_alt: cell.fill = alt_fill
            if col == "timestamp":
                cell.font = ts_font; cell.alignment = c_align
            else:
                cell.font = data_font; cell.alignment = c_align
                if isinstance(val, float): cell.number_format = "0.000"

    ws.column_dimensions["A"].width = 27
    for ci in range(2, len(sorted_cols)+1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 13
    ws.freeze_panes = "B2"

    # Export Info sheet
    ws2 = wb.create_sheet("Export Info")
    info = [
        ("Generated",     datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Total Rows",    len(records)),
        ("Total Columns", len(sorted_cols)),
        ("First Sample",  records[0].get("timestamp","")),
        ("Last Sample",   records[-1].get("timestamp","")),
        ("Poll Interval", f"{POLL_INTERVAL}s"),
        ("Commands",      "summary, stats, devs, pools"),
    ]
    for r,(k,v) in enumerate(info,1):
        ws2.cell(r,1,k).font = Font(bold=True,name="Arial",size=10)
        ws2.cell(r,2,v).font = Font(name="Arial",size=10)
    ws2.column_dimensions["A"].width=18; ws2.column_dimensions["B"].width=35

    # Column legend sheet
    ws3 = wb.create_sheet("Columns")
    for ci,hdr in enumerate(["#","Column Name","Section"],1):
        ws3.cell(1,ci,hdr).font = Font(bold=True,name="Arial")
    for i,col in enumerate(sorted_cols,1):
        so = section_order(col)[0]
        ws3.cell(i+1,1,i); ws3.cell(i+1,2,col)
        ws3.cell(i+1,3,SECTION_NAMES.get(so,""))
    ws3.column_dimensions["A"].width=6
    ws3.column_dimensions["B"].width=35
    ws3.column_dimensions["C"].width=12

    ts_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    suffix = f"_{label}" if label else ""
    fpath  = os.path.join(EXCEL_DIR, f"miner_data{suffix}_{ts_str}.xlsx")
    wb.save(fpath)
    return fpath

# ── POLLER THREAD ─────────────────────────────────────────────────────────────

class MinerPoller:
    def __init__(self, ip, port):
        self.ip=ip; self.port=port; self.records=[]
        self.lock=threading.Lock(); self._stop=threading.Event()
        self.online=False; self.last_ts="never"; self.last_err=""
        self.poll_count=0; self.error_count=0
        self._thread=threading.Thread(target=self._loop, daemon=True)

    def start(self):
        loaded = load_backup(BACKUP_FILE)
        with self.lock: self.records = loaded
        if loaded:
            print(f"{C_CYAN}  Loaded {len(loaded)} existing records from backup.{C_RESET}")
        self._thread.start()

    def stop(self): self._stop.set()
    def get_records(self):
        with self.lock: return list(self.records)

    def _loop(self):
        while not self._stop.is_set():
            t0 = time.time()
            try:
                fields = poll_all(self.ip, self.port)
                errs   = fields.pop("_errors", {})
                record = {"timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f"), **fields}
                with self.lock: self.records.append(record)
                append_backup(BACKUP_FILE, record)
                self.online=True; self.last_ts=record["timestamp"]
                self.poll_count+=1
                self.last_err = (f"partial: {errs}") if errs else ""
            except Exception as e:
                self.online=False; self.last_err=str(e); self.error_count+=1
            self._stop.wait(max(0, POLL_INTERVAL-(time.time()-t0)))

# ── CONSOLE UI ────────────────────────────────────────────────────────────────

def print_banner(ip, port):
    print(f"""
{C_BOLD}{C_CYAN}╔══════════════════════════════════════════════════╗
║       MINER DATA COLLECTOR  v2.0                 ║
║  summary · stats · devs · pools  every {POLL_INTERVAL}s       ║
╚══════════════════════════════════════════════════╝{C_RESET}
  Target : {C_YELLOW}{ip}:{port}{C_RESET}
  Backup : {BACKUP_FILE}
  Exports: {EXCEL_DIR}/
""")

def print_status(p):
    st = f"{C_GREEN}ONLINE{C_RESET}" if p.online else f"{C_RED}OFFLINE{C_RESET}"
    with p.lock: n = len(p.records)
    print(f"\n  Status      : {st}")
    print(f"  Last poll   : {p.last_ts}")
    print(f"  Records     : {C_BOLD}{n}{C_RESET}")
    print(f"  Polls OK    : {p.poll_count}")
    print(f"  Errors      : {p.error_count}")
    if p.last_err: print(f"  Last error  : {C_RED}{p.last_err}{C_RESET}")

def print_menu():
    print(f"""
{C_BOLD}  ── MENU ───────────────────────────────────────{C_RESET}
  {C_CYAN}1{C_RESET}  Export ALL data to Excel
  {C_CYAN}2{C_RESET}  Export LAST N rows to Excel
  {C_CYAN}3{C_RESET}  Show status
  {C_CYAN}4{C_RESET}  Show last reading (all fields)
  {C_CYAN}5{C_RESET}  Show discovered columns
  {C_CYAN}6{C_RESET}  Clear in-memory data (backup kept on disk)
  {C_CYAN}0{C_RESET}  Quit
  ───────────────────────────────────────────────""")

def show_last_reading(p):
    recs = p.get_records()
    if not recs: print(f"  {C_YELLOW}No data yet.{C_RESET}"); return
    rec = recs[-1]
    print(f"\n  {C_BOLD}Last reading @ {rec.get('timestamp','')}{C_RESET}")
    groups = {"SUMMARY":{},"STATS":{},"DEVS":{},"POOLS":{}}
    for k,v in rec.items():
        if k=="timestamp": continue
        if k.startswith("dev"):   groups["DEVS"][k]=v
        elif k.startswith("pool"): groups["POOLS"][k]=v
        elif section_order(k)[0]==2: groups["STATS"][k]=v
        else: groups["SUMMARY"][k]=v
    for sec, data in groups.items():
        if not data: continue
        print(f"\n  {C_BOLD}{C_CYAN}── {sec} ──{C_RESET}")
        for k,v in sorted(data.items()):
            vstr = f"{v:.4f}" if isinstance(v,float) else str(v)
            print(f"    {k:<38} {vstr}")

def show_columns(p):
    recs = p.get_records()
    if not recs: print(f"  {C_YELLOW}No data yet.{C_RESET}"); return
    all_keys = set()
    for r in recs: all_keys.update(k for k in r if not k.startswith("_"))
    cols = sorted(all_keys, key=section_order)
    print(f"\n  {C_BOLD}{len(cols)} columns discovered:{C_RESET}")
    for i,c in enumerate(cols,1):
        sec = SECTION_NAMES.get(section_order(c)[0],"")
        print(f"    {i:>3}. [{sec:<8}] {c}")

def run_console(poller):
    while True:
        print_menu()
        try:
            choice = input(f"  {C_BOLD}Choice > {C_RESET}").strip()
        except (EOFError, KeyboardInterrupt):
            choice = "0"

        if choice == "1":
            recs = poller.get_records()
            try:
                path = export_excel(recs)
                print(f"\n  {C_GREEN}Saved {len(recs)} rows to {path}{C_RESET}")
            except Exception as e:
                print(f"\n  {C_RED}Export failed: {e}{C_RESET}")
        elif choice == "2":
            try: n = int(input("  Last N rows: ").strip())
            except ValueError: print(f"  {C_RED}Invalid.{C_RESET}"); continue
            recs = poller.get_records()[-n:]
            try:
                path = export_excel(recs, label=f"last{n}")
                print(f"\n  {C_GREEN}Saved {len(recs)} rows to {path}{C_RESET}")
            except Exception as e:
                print(f"\n  {C_RED}Export failed: {e}{C_RESET}")
        elif choice == "3": print_status(poller)
        elif choice == "4": show_last_reading(poller)
        elif choice == "5": show_columns(poller)
        elif choice == "6":
            with poller.lock: poller.records.clear()
            print(f"  {C_YELLOW}In-memory cleared. Disk backup untouched.{C_RESET}")
        elif choice == "0":
            print(f"\n  {C_YELLOW}Stopping...{C_RESET}"); poller.stop()
            print(f"  {C_GREEN}Goodbye.{C_RESET}\n"); sys.exit(0)
        else:
            print(f"  {C_RED}Unknown option.{C_RESET}")

# ── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Antminer L3+ 24/7 Data Collector")
    parser.add_argument("--ip",   default="192.168.1.100", help="Miner IP")
    parser.add_argument("--port", default=4028, type=int,  help="cgminer port")
    args = parser.parse_args()

    if not EXCEL_OK:
        print(f"{C_YELLOW}WARNING: openpyxl not installed.\n"
              f"Run: C:\\Users\\User\\AppData\\Local\\Programs\\Python\\Python313\\"
              f"python.exe -m pip install openpyxl{C_RESET}\n")

    print_banner(args.ip, args.port)
    poller = MinerPoller(args.ip, args.port)
    poller.start()
    print(f"  {C_GREEN}Poller started — every {POLL_INTERVAL}s{C_RESET}")
    print(f"  {C_CYAN}First reading in ~{POLL_INTERVAL}s.{C_RESET}")
    try:
        run_console(poller)
    except KeyboardInterrupt:
        print(f"\n  {C_YELLOW}Stopping...{C_RESET}"); poller.stop(); sys.exit(0)

if __name__ == "__main__":
    main()