"""
Training API Routes
- train-now with background task + status polling
- model evaluation with severity thresholds
- model export (.pkl) for sharing
- model import from .pkl file
- data export to CSV/Excel
"""
import os, pickle, json, io, csv
from datetime import datetime, timezone
from pathlib import Path
from fastapi import APIRouter, HTTPException, Depends, BackgroundTasks, UploadFile, File
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func

from backend.db.timeseries import (get_session, get_readings_for_training,
                                    get_active_training, get_completed_training)
from backend.db.models import Miner, TrainingRun, TelemetryReading
from backend.ml.trainer import train_models, get_baseline, import_model

router    = APIRouter()
MODELS_DIR = Path(__file__).parent.parent.parent / "data" / "models"

class StartTrainingRequest(BaseModel):
    target_samples: int = 240


@router.post("/{miner_id}/start")
async def start_training(miner_id, body: StartTrainingRequest, session: AsyncSession = Depends(get_session)):
    miner = await session.get(Miner, miner_id)
    if not miner: raise HTTPException(404, f"Miner {miner_id} not found")
    active = await get_active_training(session, miner_id)
    if active: active.is_active = False; active.status = "cancelled"; await session.commit()
    run = TrainingRun(miner_id=miner_id, target_samples=body.target_samples, status="learning")
    session.add(run); await session.commit()
    return {"run_id": run.id, "status": "learning", "target_samples": body.target_samples}


@router.get("/{miner_id}/status")
async def training_status(miner_id: str, session: AsyncSession = Depends(get_session)):
    active    = await get_active_training(session, miner_id)
    completed = await get_completed_training(session, miner_id)
    count_res = await session.execute(
        select(func.count()).where(TelemetryReading.miner_id == miner_id, TelemetryReading.poll_ok == True)
    )
    sample_count = count_res.scalar_one()

    if active and active.status == "training":
        return {"phase": "training", "run_id": active.id, "sample_count": sample_count,
                "message": "Training ML models... polling every 3s"}

    if active and active.status == "learning":
        progress = min(100, int(sample_count / max(active.target_samples, 1) * 100))
        return {
            "phase": "learning", "run_id": active.id,
            "sample_count": sample_count, "target_samples": active.target_samples,
            "progress_pct": progress, "can_train_now": sample_count >= 10,
            "started_at": active.started_at.isoformat(),
        }

    if completed:
        baseline   = get_baseline(miner_id)
        evaluation = _load_evaluation(completed.model_path)
        return {
            "phase": "monitoring", "run_id": completed.id,
            "trained_at": completed.finished_at.isoformat() if completed.finished_at else None,
            "sample_count": completed.sample_count,
            "baseline_features": list(baseline.keys()) if baseline else [],
            "evaluation": evaluation,
        }

    return {"phase": "idle", "sample_count": sample_count,
            "message": "Start a learning phase to enable anomaly detection."}


@router.post("/{miner_id}/train-now")
async def train_now(miner_id: str, bg: BackgroundTasks, session: AsyncSession = Depends(get_session)):
    readings = await get_readings_for_training(session, miner_id)
    clean    = [r for r in readings if not any(k.startswith("_") for k in r)]
    if len(clean) < 10:
        raise HTTPException(400, f"Need at least 10 clean readings, got {len(clean)}")

    active = await get_active_training(session, miner_id)
    if active:
        active.status = "training"; run = active
    else:
        run = TrainingRun(miner_id=miner_id, target_samples=len(clean),
                          sample_count=len(clean), status="training")
        session.add(run)
    await session.commit()
    run_id = run.id

    async def _do():
        from backend.db.timeseries import AsyncSessionLocal
        try:
            result = train_models(miner_id, clean)
            async with AsyncSessionLocal() as s:
                r = await s.get(TrainingRun, run_id)
                if r:
                    r.status = "complete"; r.finished_at = datetime.now(timezone.utc)
                    r.sample_count = len(clean); r.baseline_stats = get_baseline(miner_id)
                    r.model_path = result.get("if_model_path"); r.is_active = False
                    await s.commit()
        except Exception as e:
            async with AsyncSessionLocal() as s:
                r = await s.get(TrainingRun, run_id)
                if r: r.status = "failed"; await s.commit()

    bg.add_task(_do)
    return {"run_id": run_id, "status": "training", "n_samples": len(clean),
            "message": "Training started. Poll /status every few seconds."}


@router.get("/{miner_id}/baseline")
async def get_baseline_stats(miner_id: str):
    b = get_baseline(miner_id)
    if not b: raise HTTPException(404, "No baseline. Run training first.")
    return b


@router.get("/{miner_id}/eda-report")
async def get_eda_report(miner_id: str):
    """Get the auto-EDA report generated during last training."""
    from backend.ml.trainer import get_eda_report
    report = get_eda_report(miner_id)
    if not report: raise HTTPException(404, "No EDA report yet. Run training first.")
    return report


@router.get("/{miner_id}/evaluate")
async def evaluate_model(miner_id: str, session: AsyncSession = Depends(get_session)):
    completed = await get_completed_training(session, miner_id)
    if not completed or not completed.model_path:
        raise HTTPException(404, "No trained model found.")
    ev = _load_evaluation(completed.model_path)
    if not ev: raise HTTPException(500, "Could not load model for evaluation.")
    baseline = get_baseline(miner_id) or {}
    ev["baseline_summary"] = {
        f: {"mean": s["mean"], "std": s["std"], "range": f"{s['p5']:.2f}–{s['p95']:.2f}"}
        for f, s in baseline.items()
    }
    ev["recommendation"] = _recommend(completed.sample_count)
    return ev


@router.get("/{miner_id}/export")
async def export_model(miner_id: str, session: AsyncSession = Depends(get_session)):
    """Download trained model as .pkl for sharing with other users."""
    completed = await get_completed_training(session, miner_id)
    if not completed or not completed.model_path:
        raise HTTPException(404, "No trained model to export.")
    if not os.path.exists(completed.model_path):
        raise HTTPException(404, "Model file missing on disk.")

    # Bundle model + baseline together for portability
    bundle_path = MODELS_DIR / f"bundle_{miner_id}.pkl"
    MODELS_DIR.mkdir(parents=True, exist_ok=True)
    baseline = get_baseline(miner_id) or {}
    with open(completed.model_path, "rb") as f:
        model_data = pickle.load(f)
    model_data["baseline"]  = baseline
    model_data["miner_id"]  = miner_id
    model_data["exported_at"] = datetime.now(timezone.utc).isoformat()
    with open(bundle_path, "wb") as f:
        pickle.dump(model_data, f)

    ts = completed.finished_at.strftime("%Y%m%d") if completed.finished_at else "unknown"
    return FileResponse(
        str(bundle_path),
        media_type="application/octet-stream",
        filename=f"miner_model_{miner_id}_{ts}.pkl",
    )


@router.post("/{miner_id}/import-model")
async def import_model_file(miner_id: str, file: UploadFile = File(...),
                             session: AsyncSession = Depends(get_session)):
    """
    Import a .pkl model file exported from another miner/user.
    The imported model's learned patterns are applied to this miner_id.
    You can then continue collecting data and retrain on top of it.
    """
    miner = await session.get(Miner, miner_id)
    if not miner: raise HTTPException(404, f"Miner {miner_id} not found")

    if not file.filename.endswith(".pkl"):
        raise HTTPException(400, "File must be a .pkl model file")

    # Save the uploaded file
    MODELS_DIR.mkdir(parents=True, exist_ok=True)
    save_path = MODELS_DIR / f"imported_{miner_id}.pkl"
    contents  = await file.read()
    with open(save_path, "wb") as f:
        f.write(contents)

    # Load and inspect
    try:
        with open(save_path, "rb") as f:
            data = pickle.load(f)
    except Exception as e:
        raise HTTPException(400, f"Invalid model file: {e}")

    baseline = data.get("baseline", {})
    ok = import_model(miner_id, str(save_path), baseline)
    if not ok:
        raise HTTPException(500, "Failed to load model into memory")

    # Create a completed training run record for this imported model
    run = TrainingRun(
        miner_id=miner_id,
        status="complete",
        finished_at=datetime.now(timezone.utc),
        sample_count=data.get("n_samples", 0),
        baseline_stats=baseline,
        model_path=str(save_path),
        is_active=False,
    )
    session.add(run); await session.commit()

    return {
        "imported": True,
        "source_miner":  data.get("miner_id", "unknown"),
        "exported_at":   data.get("exported_at", "unknown"),
        "n_features":    len(data.get("feature_names", [])),
        "features":      data.get("feature_names", []),
        "message": (
            "Model imported successfully. The miner will now use these learned patterns. "
            "You can start collecting data and retrain to adapt the model to this specific miner."
        )
    }


@router.get("/{miner_id}/export-data")
async def export_data(miner_id: str, fmt: str = "csv", session: AsyncSession = Depends(get_session)):
    """
    Export all collected telemetry data as CSV or Excel.
    fmt = 'csv' or 'excel'
    """
    readings = await get_readings_for_training(session, miner_id, max_samples=100000)
    if not readings:
        raise HTTPException(404, "No data collected yet.")

    # Collect all field names
    all_keys = set()
    for r in readings:
        all_keys.update(k for k in r.keys() if k != "timestamp" and not k.startswith("_"))
    fieldnames = ["timestamp"] + sorted(all_keys)

    if fmt == "excel":
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Telemetry_{miner_id}"
            ws.append(fieldnames)
            for r in readings:
                row = [str(r.get("timestamp", ""))] + [r.get(k, "") for k in fieldnames[1:]]
                ws.append(row)
            buf = io.BytesIO()
            wb.save(buf); buf.seek(0)
            return StreamingResponse(
                buf,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=miner_{miner_id}_data.xlsx"}
            )
        except ImportError:
            pass  # fall through to CSV

    # CSV (default or fallback)
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    for r in readings:
        clean = {"timestamp": str(r.get("timestamp", ""))}
        clean.update({k: r.get(k, "") for k in fieldnames[1:]})
        writer.writerow(clean)
    buf.seek(0)
    return StreamingResponse(
        io.BytesIO(buf.getvalue().encode()),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=miner_{miner_id}_data.csv"}
    )


@router.post("/{miner_id}/import-data")
async def import_training_data(
    miner_id: str,
    file: UploadFile = File(...),
    session: AsyncSession = Depends(get_session),
):
    """
    Import telemetry data from a CSV or Excel file to use as training data.
    The file must have a 'timestamp' column plus numeric feature columns.
    This lets you set a custom starting point or import data from another source.
    """
    miner = await session.get(Miner, miner_id)
    if not miner:
        raise HTTPException(404, f"Miner {miner_id} not found")

    fname = file.filename or ""
    contents = await file.read()

    rows = []
    try:
        if fname.endswith((".xlsx", ".xls")):
            import openpyxl, io as _io
            wb = openpyxl.load_workbook(_io.BytesIO(contents), read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
        elif fname.endswith(".csv"):
            import io as _io, csv as _csv
            text = contents.decode(errors="replace")
            reader = _csv.DictReader(_io.StringIO(text))
            rows = list(reader)
        else:
            raise HTTPException(400, "File must be .csv, .xlsx, or .xls")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"Could not read file: {e}")

    if not rows:
        raise HTTPException(400, "File is empty or has no data rows")

    # Parse rows into TelemetryReading records
    inserted = 0
    skipped  = 0
    ts_now   = datetime.now(timezone.utc)

    for i, row in enumerate(rows):
        values = {}
        for k, v in row.items():
            if k in ("timestamp", "Timestamp", "time", "Time", ""): continue
            if v is None or v == "": continue
            try:
                values[str(k)] = float(str(v).strip())
            except (ValueError, TypeError):
                pass

        if not values:
            skipped += 1; continue

        # Parse timestamp if present
        ts_col = row.get("timestamp") or row.get("Timestamp") or row.get("time")
        try:
            from datetime import datetime as dt
            ts = dt.fromisoformat(str(ts_col).replace("Z", "+00:00"))
            if ts.tzinfo is None:
                from datetime import timezone as tz
                ts = ts.replace(tzinfo=tz.utc)
        except Exception:
            # Use evenly-spaced synthetic timestamps
            from datetime import timedelta
            ts = ts_now - timedelta(seconds=30 * (len(rows) - i))

        from backend.db.models import TelemetryReading as TR
        session.add(TR(miner_id=miner_id, timestamp=ts, values=values, poll_ok=True))
        inserted += 1

    await session.commit()

    return {
        "imported": inserted,
        "skipped":  skipped,
        "total_rows": len(rows),
        "message": (
            f"Imported {inserted} readings from {fname}. "
            f"You can now train the model using this data."
        )
    }


def _load_evaluation(model_path: str) -> dict:
    if not model_path or not os.path.exists(model_path):
        return None
    try:
        with open(model_path, "rb") as f:
            d = pickle.load(f)
        return {
            "model_type":        "Isolation Forest",
            "n_features":        len(d.get("feature_names", [])),
            "features":          d.get("feature_names", []),
            "threshold_yellow":  round(d.get("threshold_yellow", d.get("threshold", 0.5)), 4),
            "threshold_red":     round(d.get("threshold_red", 0.75), 4),
        }
    except:
        return None


def _recommend(n_samples: int) -> str:
    if n_samples < 30:   return "Very few samples — collect 2+ hours and retrain"
    if n_samples < 120:  return "Functional but improve with 4+ hours of data"
    if n_samples < 480:  return "Good model — retrain only after hardware/config changes"
    return "Excellent — retrain only after major changes (firmware, hardware repair)"


@router.post("/{miner_id}/simulate")
async def simulate_live_feed(
    miner_id: str,
    file: UploadFile = File(...),
    bg: BackgroundTasks = None,
    session: AsyncSession = Depends(get_session),
):
    """
    Feed CSV/Excel data row-by-row as if it were coming from a live miner.
    Each row is inserted with a 1-second delay and triggers the full
    ML pipeline (scoring, rules, WS broadcast) exactly like real polling.
    
    Use this to:
    - Test the system without a physical miner connected
    - Replay historical data through the anomaly detector
    - Demonstrate the dashboard with realistic data flow
    """
    miner = await session.get(Miner, miner_id)
    if not miner:
        raise HTTPException(404, f"Miner {miner_id} not found")

    fname = file.filename or ""
    contents = await file.read()

    rows = []
    try:
        if fname.endswith((".xlsx", ".xls")):
            import openpyxl, io as _io
            wb = openpyxl.load_workbook(_io.BytesIO(contents), read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
        elif fname.endswith(".csv"):
            import io as _io, csv as _csv
            text = contents.decode(errors="replace")
            reader = _csv.DictReader(_io.StringIO(text))
            rows = list(reader)
        else:
            raise HTTPException(400, "File must be .csv or .xlsx")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"Could not read file: {e}")

    if not rows:
        raise HTTPException(400, "File is empty")

    # Parse all rows into numeric dicts
    parsed_rows = []
    for row in rows:
        values = {}
        for k, v in row.items():
            if k in ("timestamp", "Timestamp", "time", "anomaly_label", "anomaly_type", ""):
                continue
            if v is None or v == "": continue
            try:
                values[str(k)] = float(str(v).strip())
            except (ValueError, TypeError):
                pass
        if values:
            parsed_rows.append(values)

    if not parsed_rows:
        raise HTTPException(400, "No numeric data found in file")

    # Run simulation in background — feeds rows through the WS pipeline
    async def _simulate():
        import asyncio
        from backend.api.routes.ws import on_new_reading

        for i, values in enumerate(parsed_rows):
            ts = datetime.now(timezone.utc)
            try:
                await on_new_reading(miner_id, ts, values)
            except Exception as e:
                import logging
                logging.getLogger("simulate").warning(f"Row {i}: {e}")
            # 0.5s delay between rows — fast enough to demo, slow enough to watch
            await asyncio.sleep(0.5)

    import asyncio
    asyncio.create_task(_simulate())

    return {
        "simulating": True,
        "total_rows": len(parsed_rows),
        "delay_seconds": 0.5,
        "estimated_duration_seconds": len(parsed_rows) * 0.5,
        "message": (
            f"Simulating {len(parsed_rows)} readings at 0.5s intervals. "
            f"Watch the dashboard — data will appear in real time. "
            f"Estimated time: {len(parsed_rows) * 0.5:.0f}s."
        ),
    }
